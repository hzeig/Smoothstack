import logging as lg
import os
import datetime
import openpyxl

# defining errors and exceptions
class Error(Exception):
    """Base class for other exceptions"""
    pass

class DateMissingError(Error):
    """Raised when date not found."""
    pass

class DataMissingError(Error):
    """Raised when some part of data not found."""

# data collection and report formatting function
def vocReport(path, date_obj):
    
    wb = openpyxl.load_workbook(path) 
    lg.info("Opened file workbook.")

    ws = wb["VOC Rolling MoM"]
    lg.info("Opened worksheet 'VOC Rolling Mom'.")
    
    lg.info("Gathering report data.")
    date = "{}, {}".format(date_obj[0], date_obj[1]) 
    month = date_obj[0]
    
    data = []
    columns = ws.iter_cols(min_row=1, max_row=1, min_col=2, max_col=13)
    
    try:
        lg.info("Locating data row for {}.".format(date))
        for column in columns:
            for cell in column:
                # check if date exists as datetime object
                if isinstance(cell.value, datetime.date):
                    cell_date = cell.value.strftime("%B, %Y")
                    lg.debug("Comparing datetime {} to {}".format(cell_date, date.capitalize()))
                    if cell_date.lower() == date:
                        col = cell.column
                        lg.info("{} data found in column {} of file.".format(date.capitalize(), col))
                        break
                # else:
                #     lg.warning(DataMissingError("Date Missing Error: Unable to find {} in datetime format.".format(date)))
                else: 
                    # if no matching datetime object, search by month name
                    string = str(cell.value)
                    lg.debug("Comparing string {} to {}".format(string, month.capitalize()))
                    if string.lower() == month:
                        col = cell.column
                    # else:
                    #     lg.error(DateMissingError("Date Missing Error: Unable to find {} in string format.".format(date)))

    except DateMissingError:
        lg.error("DateMissingError: Date from file name not found in file.")


    try:
        lg.info("Collecting data.")
        column_data = ws.iter_cols(min_row=4, max_row=9, min_col=col, max_col=col)
        for cells in column_data:
            for cell in cells:
                if cell.value > 1:
                    data.append(cell.value)
                else:
                    pass
        lg.info("Data collected.")

        lg.info("Analyzing data.")
        status = []
        for value in data:
            if value.index == 1:
                if value > 200:
                    status.append('good (>200)')
                else:
                    status.append('bad (<200)')
            else:
                if value > 100:
                    status.append('good (>100)')
                else:
                    status.append('bad (<100)')

        zipped_data = list(zip(data, status))

        report = "VOC Report {}: \n\
                Promoters: {}, \n\
                Passives: {}, \n\
                Decractors: {}, "\
                .format(cell_date, *list(map(lambda x: "{}, {}".format(x[0], x[1]), zipped_data)))
        lg.info("Data successfully gathered.")
        lg.debug(report)
        
        return report

    except DataMissingError:
        lg.error("DataMissingError: Data for file date not found.")